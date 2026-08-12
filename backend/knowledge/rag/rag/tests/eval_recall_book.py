"""真实书籍召回率测试(第三次 RAG 测试):以《UNIX网络编程 卷1:套接字联网API》第3版中文版
前 5000 字为语料。

与前两次测试的区别:
  第一次 test(eval_recall.py):官方 200 条词库 + 人工造的查询。
  第二次 test(eval_recall_article.py):ICT 会议论文全文 + 文章实际表面形式。
  第三次 test(eval_recall_book.py):网络编程经典书前 5000 字(扫描版 PDF OCR,
    第 16~20 页·第 1 章开头)+ 书内实际表面形式。领域从「ICT 综合」收窄到「计算机网络」。

测试集生成逻辑(与 eval_recall_article.py 同口径):
  1) 载入官方 200 条术语(ICT_Terms_200.xlsx);
  2) 对每条术语解析出「全称 / 缩写」两种表面形式;
  3) 在书前 5000 字(规范化:连字、空白)中检测哪些形式真实出现;
  4) 命中查询 A = 出现形式的原文直查 / 缩写直查(预期 = 词库术语);
  5) 命中查询 B = 书中高频出现的「词库词条子形式的网络缩写」(如 TCP、IP,
     是 TCP/IP 词条缩写 TCP/IP 的子形式,单独出现时也应召回父词条);
  6) 命中查询 C = 网络领域语义改写(模拟译前术语提取 LLM 输出的同义改写);
  7) 命中查询 D = 中文反向查询(产品支持中文源文,中文候选经别名层命中英文词条);
  8) 负例 E   = 书中出现但词库没有的网络术语(预期 = 未命中)。

用法:
    python tests/eval_recall_book.py                # 输出测试集 + 结果
    python tests/eval_recall_book.py --json-only    # 只生成测试集 JSON
"""
import argparse
import json
import re
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook                      # noqa: E402

from knowledge_base import config, init_collection, search_rag   # noqa: E402
from knowledge_base.embedder import embed                        # noqa: E402
from knowledge_base.glossary_split import split_combined_term    # noqa: E402
from knowledge_base.rag_aliases import lookup_alias              # noqa: E402
from knowledge_base.rag_terms import _build_where, _collection, _parse_results  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parent.parent
GLOSSARY_XLSX = PROJECT_ROOT / "ICT_Terms_200.xlsx"
BOOK_TXT = PROJECT_ROOT / "data" / "unix_book_first5000.txt"
OUT_QUERIES = Path(__file__).resolve().parent / "eval_queries_book.json"
TOP_K = 5

# 复用第二次测试的规范化与表面形式解析逻辑
from eval_recall_article import (                            # noqa: E402
    term_patterns, compile_pattern, norm_text_raw, norm_text,
)

# ── 第三次测试专项:书内高频网络缩写(TCP/IP 词条拆分后的子词条缩写)────────
# 这些缩写全书反复出现,单独作为术语候选查询时经别名层命中拆分后的子词条:
#   TCP → Transmission Control Protocol (TCP),IP → Internet Protocol (IP)。
# 词条自身表面形式解析(TCP/IP)覆盖不到「裸 TCP / 裸 IP」,故单独列出考验别名子形式。
BOOK_ABBR = [
    {"query": "TCP", "expected": "Transmission Control Protocol (TCP)",
     "kind": "书内缩写", "note": "书内高频缩写-应命中"},
    {"query": "IP", "expected": "Internet Protocol (IP)",
     "kind": "书内缩写", "note": "书内高频缩写-应命中"},
]

# ── 网络领域语义改写(模拟 LLM 同义改写,考验语义路径)──────────────────
REWRITE_QUERIES = [
    {"query": "virtual private network", "expected": "Virtual Private Network (VPN)",
     "kind": "改写", "note": "网络领域改写-应命中"},
    {"query": "software-defined networking", "expected": "Software-Defined Networking (SDN)",
     "kind": "改写", "note": "网络领域改写-应命中"},
    {"query": "packet switching", "expected": "Packet Switching",
     "kind": "改写", "note": "网络领域改写-应命中"},
    {"query": "load balancing", "expected": "Load Balancing",
     "kind": "改写", "note": "网络领域改写-应命中"},
    {"query": "firewall protection", "expected": "Firewall",
     "kind": "改写", "note": "网络领域改写-应命中"},
    {"query": "domain name system", "expected": "Domain Name System (DNS)",
     "kind": "改写", "note": "网络领域改写-应命中"},
    {"query": "wireless local area network", "expected": "Wireless Local Area Network (WLAN)",
     "kind": "改写", "note": "网络领域改写-应命中"},
    {"query": "metropolitan area network", "expected": "Metropolitan Area Network (MAN)",
     "kind": "改写", "note": "网络领域改写-应命中"},
    {"query": "software-defined wide area network", "expected": "Software-Defined Wide Area Network (SD-WAN)",
     "kind": "改写", "note": "网络领域改写-应命中"},
    {"query": "mobile fifth generation networks", "expected": "5G (Fifth Generation Mobile Network)",
     "kind": "改写", "note": "网络领域改写-应命中"},
]

# ── 中文反向查询(产品支持中文源文,中文候选经别名层命中英文词条)────────
ZH_REVERSE = [
    {"query": "传输控制协议", "expected": "Transmission Control Protocol (TCP)",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "网际协议", "expected": "Internet Protocol (IP)",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "局域网", "expected": "Local Area Network (LAN)",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "广域网", "expected": "Wide Area Network (WAN)",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "路由器", "expected": "Router",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "应用程序编程接口", "expected": "Application Programming Interface (API)",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "域名系统", "expected": "Domain Name System (DNS)",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "防火墙", "expected": "Firewall",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "负载均衡", "expected": "Load Balancing",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "分组交换", "expected": "Packet Switching",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "虚拟专用网络", "expected": "Virtual Private Network (VPN)",
     "kind": "中文反向", "note": "中文源文-应命中"},
    {"query": "软件定义网络", "expected": "Software-Defined Networking (SDN)",
     "kind": "中文反向", "note": "中文源文-应命中"},
]

# ── 负例:书中出现但词库没有的网络术语(应未命中)────────────────────
NEGATIVES = [
    "UDP", "IPv4", "IPv6", "HTTP", "SMTP", "RPC", "SCTP", "socket",
    "daemon", "POSIX", "MTU", "MSS", "PDU", "client", "server",
    "fragment", "intranet", "sockaddr",
]


def load_glossary() -> list[dict]:
    wb = load_workbook(GLOSSARY_XLSX, read_only=True, data_only=True)
    ws = wb.active
    terms = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        # 缝合怪词条按拆分后的独立词条建测试词汇,保证 expected 与库内词条名一致
        for st, strans in split_combined_term(str(row[1]).strip(), str(row[2]).strip()):
            terms.append({"term": st, "translation": strans})
    return terms


def raw_query(query: str, domain: str = "") -> list[dict]:
    col = _collection()
    if col.count() == 0:
        return []
    res = col.query(query_embeddings=[embed(query)],
                    where=_build_where(domain, config.DEFAULT_USER),
                    n_results=col.count())
    return _parse_results(res)


def main() -> None:
    ap = argparse.ArgumentParser(description="真实书籍(UNIX网络编程前5000字)召回率测试")
    ap.add_argument("--json-only", action="store_true", help="只生成测试集 JSON,不跑评估")
    args = ap.parse_args()

    # 1. 载入词库 + 书前5000字
    glossary = load_glossary()
    raw_src = BOOK_TXT.read_text(encoding="utf-8")
    book_raw = norm_text_raw(raw_src)
    book_low = norm_text(raw_src)
    print(f"词库 {len(glossary)} 条 | 书前5000字 {len(book_low)} 字符")

    # 2. 检测每条术语在书中的表面形式
    term_surfaces: list[dict] = []
    for g in glossary:
        found: list[dict] = []
        for p, is_abbr in term_patterns(g["term"]):
            if is_abbr:
                pat = compile_pattern(p, True)
                ok = pat.search(book_raw)
            else:
                pat = compile_pattern(p, False)
                ok = pat.search(book_low)
            if ok:
                found.append({"text": p, "kind": "缩写" if is_abbr else "全称"})
        seen = set()
        unique = []
        for f in found:
            if f["text"] not in seen:
                seen.add(f["text"])
                unique.append(f)
        term_surfaces.append({**g, "surfaces": unique})

    present = [t for t in term_surfaces if t["surfaces"]]
    print(f"书中出现词库术语: {len(present)}/{len(glossary)}")

    # 3. 命中查询 A:每个出现表面形式一条(优先缩写)
    hit_queries: list[dict] = []
    seen_q = set()
    for t in present:
        for f in sorted(t["surfaces"], key=lambda x: 0 if x["kind"] == "缩写" else 1):
            key = (f["text"].lower(), t["term"])
            if key in seen_q:
                continue
            seen_q.add(key)
            hit_queries.append({
                "query": f["text"],
                "expected": t["term"],
                "translation": t["translation"],
                "kind": f["kind"],
                "note": f"书内表面形式-{f['kind']}",
            })

    queries = hit_queries + BOOK_ABBR + REWRITE_QUERIES + ZH_REVERSE
    miss_queries = [{"query": q, "expected": "", "note": "书中出现但词库无-应未命中"}
                    for q in NEGATIVES]
    queries += miss_queries

    with open(OUT_QUERIES, "w", encoding="utf-8") as f:
        json.dump(queries, f, ensure_ascii=False, indent=2)
    n_hit = sum(1 for q in queries if q.get("expected"))
    print(f"测试集已写入 {OUT_QUERIES.name}: 命中 {n_hit} 条 / 负例 {len(miss_queries)} 条")

    if args.json_only:
        return

    # 4. 评估(口径与 eval_recall_article.py 一致)
    n = len(queries)
    r1 = r5 = mrr_sum = judge_ok = 0
    alias_total = sem_total = alias_top1 = sem_top1 = 0
    hit_sims: list[float] = []
    miss_max_sims: list[float] = []
    fail_rows = []
    fp_rows = []

    print("=" * 100)
    print(f"真实书籍召回率测试 | 查询 {n} 条(命中 {n_hit} / 负例 {n - n_hit}) | "
          f"当前阈值 {config.RAG_MIN_SIMILARITY}")
    print("=" * 100)

    for q in queries:
        query = q["query"]
        expected = q["expected"]
        note = q.get("note", "")
        t0 = time.time()
        raw = raw_query(query)
        results = search_rag(query, top_k=TOP_K)
        dt = time.time() - t0

        found = [r.term for r in results]
        rank = found.index(expected) + 1 if expected in found else 0
        exp_sim = 0.0
        for e in raw:
            if e.get("term") == expected:
                exp_sim = float(e.get("_similarity", 0.0))
                break
        top1 = results[0] if results else None
        top1_raw = raw[0] if raw else None
        top1_sim = float(top1_raw.get("_similarity", 0.0)) if top1_raw else 0.0
        top1_tr = top1.translation if top1 else ""

        expect_hit = bool(expected)
        decided_hit = len(results) > 0
        ok = (decided_hit == expect_hit)
        if ok:
            judge_ok += 1

        if expect_hit:
            hit_sims.append(exp_sim)
            al = lookup_alias(query, "", "")
            is_alias = bool(al) and al[0] == expected
            if is_alias:
                alias_total += 1
                if rank == 1:
                    alias_top1 += 1
            else:
                sem_total += 1
                if rank == 1:
                    sem_top1 += 1
            if rank == 1:
                r1 += 1; r5 += 1; mrr_sum += 1.0
            elif rank > 0:
                r5 += 1; mrr_sum += 1.0 / rank
        else:
            miss_max_sims.append(top1_sim)
            if decided_hit:
                fp_rows.append((query, top1.term if top1 else "", top1_tr, top1_sim))

        mark = "PASS" if ok else "FAIL"
        print(f"[{mark}] {query!r:<46} rank={rank or '-':<2} "
              f"top1={top1.term if top1 else '∅'}({top1_sim:.2f}) "
              f"| {note} | {dt*1000:.0f}ms")
        if not ok and expect_hit:
            fail_rows.append((query, expected, top1.term if top1 else "", exp_sim))

    recall1 = r1 / n_hit * 100 if n_hit else 0.0
    recall5 = r5 / n_hit * 100 if n_hit else 0.0
    mrr = mrr_sum / n_hit if n_hit else 0.0
    judge_acc = judge_ok / n * 100 if n else 0.0

    print("=" * 100)
    print(f"Recall@1 = {r1}/{n_hit} = {recall1:.1f}%")
    print(f"Recall@5 = {r5}/{n_hit} = {recall5:.1f}%")
    print(f"MRR      = {mrr:.3f}")
    print(f"命中判定准确率 = {judge_ok}/{n} = {judge_acc:.1f}%")

    print("-" * 100)
    print("路径分类统计(按 search_rag 实际命中路径):")
    if alias_total:
        print(f"  别名路径(查字典·确定性): {alias_total}/{n_hit} 条查询 "
              f"→ top1 命中 {alias_top1}(召回率 {alias_top1/alias_total*100:.1f}%)")
    else:
        print("  别名路径: 0 条查询(测试集无别名可命中的表面形式)")
    if sem_total:
        print(f"  语义路径(近似·≥{config.RAG_MIN_SIMILARITY:.2f}): {sem_total}/{n_hit} 条查询 "
              f"→ top1 命中 {sem_top1}(召回率 {sem_top1/sem_total*100:.1f}%)")
    else:
        print(f"  语义路径: 0 条查询(所有查询均被别名层覆盖,语义路径未被本次测试集触及)")
    print(f"  合计: {n_hit} 条查询 → top1 命中 {r1}(Recall@1 = {recall1:.1f}%)")
    print("  ── 说明:别名路径是「查字典」,100% 准确但只能覆盖词库已有形式;"
          "语义路径是「近似」,处理词库未精确覆盖的查询(改写/子形式/库外),能力见下方阈值敏感性表")

    print("-" * 100)
    print("阈值敏感性(0.60~0.85,仅语义能力评估;别名命中的查询不受阈值影响):")
    for t in [0.60, 0.65, 0.70, 0.75, 0.80, 0.85]:
        keep = sum(1 for s in hit_sims if s >= t)
        fp = sum(1 for s in miss_max_sims if s >= t)
        print(f"  阈值 ≥ {t:.2f}: 命中保留 {keep}/{n_hit} "
              f"(召回率 {keep/n_hit*100 if n_hit else 0:.0f}%)  |  负例误判 {fp}/{n-n_hit}")
    if hit_sims:
        srt = sorted(hit_sims)
        print(f"  命中查询相似度: min={srt[0]:.2f} max={srt[-1]:.2f} median={srt[n_hit//2]:.2f}")

    if fail_rows:
        print("-" * 100)
        print("命中类未通过项:")
        for query, expected, top1, sim in fail_rows:
            print(f"  {query!r}: 期望 {expected!r} / 实际 top1 {top1!r} / 期望相似度 {sim:.2f}")
    if fp_rows:
        print("-" * 100)
        print("负例误报项(search_rag 返回了非空):")
        for query, top1, tr, sim in fp_rows:
            print(f"  {query!r}: → {top1!r} 译={tr!r} sim={sim:.2f}")

    # 书覆盖度
    print("-" * 100)
    print(f"书覆盖度:前5000字出现词库术语 {len(present)}/{len(glossary)} "
          f"({len(present)/len(glossary)*100:.1f}%)")
    print("出现术语:", " | ".join(t["term"] for t in present))


if __name__ == "__main__":
    main()

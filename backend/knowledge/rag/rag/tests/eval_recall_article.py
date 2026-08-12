"""真实文档召回率测试:以 `ICT analysis and applicantions(1).pdf` 为语料。

与前次官方词库测试的区别:查询不再是人工造的 35 条,而是「文章中实际出现的
术语表面形式」——模拟译前术语提取第一层扫描真实文档时,RAG 能否召回
词库中存在的正确术语,以及文章里出现但词库没有的 ICT 术语会不会被误报。

测试集生成逻辑:
  1) 载入官方 200 条术语(ICT_Terms_200.xlsx);
  2) 对每条术语解析出「全称 / 缩写」两种表面形式;
  3) 在文章全文(规范化:连字、空白)中检测哪些形式真实出现;
  4) 命中查询 = 出现形式的原文直查 / 缩写直查(预期 = 词库术语);
  5) 负例     = 文章中出现但词库没有的 ICT 术语(预期 = 未命中)。

用法:
    python tests/eval_recall_article.py                # 输出测试集 + 结果
    python tests/eval_recall_article.py --json-only    # 只生成测试集 JSON
"""
import argparse
import json
import re
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook                      # noqa: E402

from knowledge_base import config, init_collection, search_rag   # noqa: E402
from knowledge_base.embedder import embed                        # noqa: E402
from knowledge_base.glossary_split import split_combined_term, split_variants  # noqa: E402
from knowledge_base.rag_aliases import lookup_alias              # noqa: E402
from knowledge_base.rag_terms import _build_where, _collection, _parse_results  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parent.parent
GLOSSARY_XLSX = PROJECT_ROOT / "ICT_Terms_200.xlsx"
ARTICLE_TXT = PROJECT_ROOT / "data" / "ict_article_fulltext.txt"
OUT_QUERIES = Path(__file__).resolve().parent / "eval_queries_article.json"
TOP_K = 5

# ── 规范化 ─────────────────────────────────────────────
_LIGATURES = {"ﬁ": "fi", "ﬂ": "fl", "ﬀ": "ff", "ﬃ": "ffi", "ﬄ": "ffl"}
_WORD_END = r"(?<![A-Za-z0-9])"
_WORD_AFTER = r"(?![A-Za-z0-9])"


def _fix_ligatures(t: str) -> str:
    for a, b in _LIGATURES.items():
        t = t.replace(a, b)
    return re.sub(r"\s+", " ", t)


def norm_text(t: str) -> str:
    """小写规范化:用于全称(忽略大小写)匹配。"""
    return _fix_ligatures(t).lower()


def norm_text_raw(t: str) -> str:
    """保留大小写规范化:用于缩写(大小写敏感)匹配。"""
    return _fix_ligatures(t)


def dehyphenate(t: str) -> str:
    """PDF 行尾断词(如 `automa- tion`)拼回原词。"""
    return re.sub(r"([A-Za-z])- ([A-Za-z])", r"\1\2", t)


def _norm_key(s: str) -> str:
    """归一化键:小写、去空格/标点,用于比较表面形式是否等价于缩写。"""
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _term_abbr_key(term: str) -> str | None:
    """取术语括号内的缩写(仅当括号在末尾、且括号内容明显为缩写时)。"""
    s = term.strip()
    m = re.match(r"^(?P<lead>.+?)\s*\((?P<inner>[^()]+)\)\s*$", s)
    if not m:
        return None
    inner = m.group("inner").strip()
    lead = m.group("lead").strip()
    # 缩写 = 括号内容,且它比全称"短"(如 CPU、AI、IoT、5G)
    if len(_norm_key(inner)) <= 12 and len(_norm_key(inner)) < len(_norm_key(lead)):
        return _norm_key(inner)
    return None


def term_patterns(term: str) -> list[tuple[str, bool]]:
    """从词库术语解析候选表面形式 → [(pattern, is_abbr)],去重。

    is_abbr 判定:该表面形式(归一化后)恰好等于术语括号内的缩写。
    全称形式统一忽略大小写匹配,缩写形式统一大小写敏感匹配。
    """
    pats: list[str] = []
    s = term.strip()
    abbr_key = _term_abbr_key(term)
    # 括号在末尾:"Central Processing Unit (CPU)" / "BIOS (Basic Input/Output System)"
    m = re.match(r"^(?P<lead>.+?)\s*\((?P<inner>[^()]+)\)\s*$", s)
    if m:
        pats += [m.group("lead").strip(), m.group("inner").strip()]
    elif "(" in s:  # 括号在中间:"Open Systems Interconnection (OSI) Model"
        pats.append(re.sub(r"\s*\([^()]*\)\s*", " ", s).strip())
        pats.append(re.sub(r"^.*?\(([^()]*)\).*?$", r"\1", s).strip())
    else:
        pats.append(s)
    expanded: list[str] = []
    for p in pats:
        expanded += split_variants(p)
    out: list[tuple[str, bool]] = []
    seen = set()
    for p in expanded:
        if not p:
            continue
        for v in (p, p.replace(" ", "")):
            if v in seen:
                continue
            seen.add(v)
            is_abbr = bool(abbr_key) and _norm_key(v) == abbr_key
            out.append((v, is_abbr))
    return out


def compile_pattern(p: str, abbr: bool) -> re.Pattern:
    """缩写:原文大小写敏感(防 Dr. 误命中 DR);全称:忽略大小写。"""
    flags = 0 if abbr else re.IGNORECASE
    return re.compile(_WORD_END + re.escape(p) + _WORD_AFTER, flags)


def load_glossary() -> list[dict]:
    wb = load_workbook(GLOSSARY_XLSX, read_only=True, data_only=True)
    ws = wb.active
    terms = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        # 缝合怪词条按拆分后的独立词条建测试词汇,保证 expected 与库内词条名一致
        for st, strans in split_combined_term(str(row[1]).strip(), str(row[2]).strip()):
            terms.append({"term": st, "translation": strans})
    return terms


def raw_query(query: str, domain: str = "") -> list[dict]:
    col = _collection()
    if col.count() == 0:
        return []
    res = col.query(query_embeddings=[embed(query)],
                    where=_build_where(domain, config.DEFAULT_USER),
                    n_results=col.count())
    return _parse_results(res)


def main() -> None:
    ap = argparse.ArgumentParser(description="真实文档(ICT 会议论文)召回率测试")
    ap.add_argument("--json-only", action="store_true", help="只生成测试集 JSON,不跑评估")
    args = ap.parse_args()

    # 1. 载入词库 + 文章
    glossary = load_glossary()
    raw_src = ARTICLE_TXT.read_text(encoding="utf-8")
    article_raw = norm_text_raw(raw_src)
    article_low = norm_text(raw_src)
    article_raw_join = dehyphenate(article_raw)
    article_low_join = dehyphenate(article_low)
    print(f"词库 {len(glossary)} 条 | 文章 {len(article_low)} 字符")

    # 2. 检测每条术语在文章中的表面形式
    term_surfaces: list[dict] = []   # {term, translation, surfaces:[{text, kind}]}
    for g in glossary:
        found: list[dict] = []
        for p, is_abbr in term_patterns(g["term"]):
            # 缩写走原文(大小写敏感);全称走小写文本(忽略大小写)
            if is_abbr:
                pat = compile_pattern(p, True)
                ok = pat.search(article_raw) or pat.search(article_raw_join)
            else:
                pat = compile_pattern(p, False)
                ok = pat.search(article_low) or pat.search(article_low_join)
            if ok:
                found.append({"text": p, "kind": "缩写" if is_abbr else "全称"})
        # 去重
        seen = set()
        unique = []
        for f in found:
            if f["text"] not in seen:
                seen.add(f["text"])
                unique.append(f)
        term_surfaces.append({**g, "surfaces": unique})

    present = [t for t in term_surfaces if t["surfaces"]]
    print(f"文章中出现词库术语: {len(present)}/{len(glossary)}")

    # 3. 构建命中查询:每个出现表面形式一条
    hit_queries: list[dict] = []
    seen_q = set()
    for t in present:
        # 优先用缩写做查询(更难、考验别名层);否则用全称
        for f in sorted(t["surfaces"], key=lambda x: 0 if x["kind"] == "缩写" else 1):
            key = (f["text"].lower(), t["term"])
            if key in seen_q:
                continue
            seen_q.add(key)
            hit_queries.append({
                "query": f["text"],
                "expected": t["term"],
                "translation": t["translation"],
                "kind": f["kind"],
                "note": f"文章表面形式-{f['kind']}",
            })

    # 3.5 语义改写查询:模拟译前术语提取 LLM 输出的同义改写(如 "cloud infrastructure"
    #      → Cloud Computing)。此类查询的归一化键不在别名表里,别名层覆盖不了,
    #      只能靠语义检索命中——让「语义路径」有样本可测,不再被别名层垄断
    #      (2026-08-10 优化日志-06 落地)。
    REWRITE_QUERIES = [
        {"query": "artificial intelligence techniques", "expected": "Artificial Intelligence (AI)"},
        {"query": "machine learning methods", "expected": "Machine Learning (ML)"},
        {"query": "deep neural networks", "expected": "Deep Learning (DL)"},
        {"query": "cloud infrastructure", "expected": "Cloud Computing"},
        {"query": "cloud services", "expected": "Cloud Computing"},
        {"query": "internet of things devices", "expected": "Internet of Things (IoT)"},
        {"query": "edge nodes", "expected": "Edge Computing"},
        {"query": "intrusion detection", "expected": "Intrusion Detection System (IDS)"},
        {"query": "data analytics techniques", "expected": "Data Analytics"},
        {"query": "serverless functions", "expected": "Serverless Computing"},
    ]
    hit_queries += [
        {**r, "kind": "改写", "translation": "", "note": "语义改写-应命中"}
        for r in REWRITE_QUERIES
    ]

    # 3.75 中文反向查询:产品支持中文源文文档(2026-08-10 确认),术语提取拿中文
    #      候选(如"数字签名")查 RAG。中文键经别名层确定性命中英文术语
    #      (rag_aliases._zh_keys),从词库动态生成前 10 条含中文译法的查询。
    ZH_REVERSE = []
    seen_zh = set()
    for g in glossary:
        tr = (g.get("translation") or "").strip()
        if not tr or not re.search(r"[一-鿿]", tr):
            continue
        if tr == g["term"]:          # 不译术语(term==译文)跳过
            continue
        m_zh = re.match(r"^(?P<lead>.+?)\s*\((?P<inner>[^()]+)\)\s*$", tr)
        q = m_zh.group("lead").strip() if m_zh else tr
        if q.lower() in seen_zh:
            continue
        seen_zh.add(q.lower())
        ZH_REVERSE.append({"query": q, "expected": g["term"],
                           "translation": "", "kind": "中文反向",
                           "note": "中文源文-应命中"})
    hit_queries += ZH_REVERSE[:10]

    # 4. 负例:文章出现但词库没有的 ICT 术语(应未命中)
    #    [2026-08-10 第一次test优化·扩库] GAN/Transfer Learning/Signature Verification/
    #    Handwritten Signature 已补入内置库(ICT_Terms_ext.csv 批次1),不再是负例;
    #    剩 12 个为当前真实未覆盖术语。
    negatives = [
        "VGG-16", "Max-Pooling", "Softmax", "Rectified Linear Unit (ReLU)",
        "Feature Extraction", "Fraud Detection", "Kubernetes", "Hadoop",
        "Apache Spark", "E-commerce", "GDPR", "V2X",
    ]
    miss_queries = [{"query": q, "expected": "", "note": "文章出现但词库无-应未命中"}
                    for q in negatives]

    queries = hit_queries + miss_queries
    with open(OUT_QUERIES, "w", encoding="utf-8") as f:
        json.dump(queries, f, ensure_ascii=False, indent=2)
    print(f"测试集已写入 {OUT_QUERIES.name}: 命中 {len(hit_queries)} 条 / 负例 {len(miss_queries)} 条")

    if args.json_only:
        return

    # 5. 评估(口径与 eval_recall.py 一致:按 search_rag 实际返回统计)
    n = len(queries)
    hit_qs = [q for q in queries if q.get("expected")]
    n_hit = len(hit_qs)
    r1 = r5 = mrr_sum = judge_ok = 0
    # 路径分类统计:别名层(查字典,确定性)vs 语义检索(近似,≥阈值)
    alias_total = sem_total = alias_top1 = sem_top1 = 0
    hit_sims: list[float] = []
    miss_max_sims: list[float] = []
    fail_rows = []
    fp_rows = []

    print("=" * 100)
    print(f"真实文档召回率测试 | 查询 {n} 条(命中 {n_hit} / 负例 {n - n_hit}) | "
          f"当前阈值 {config.RAG_MIN_SIMILARITY}")
    print("=" * 100)

    for q in queries:
        query = q["query"]
        expected = q["expected"]
        note = q.get("note", "")
        t0 = time.time()
        raw = raw_query(query)
        results = search_rag(query, top_k=TOP_K)
        dt = time.time() - t0

        found = [r.term for r in results]
        rank = found.index(expected) + 1 if expected in found else 0
        exp_sim = 0.0
        for e in raw:
            if e.get("term") == expected:
                exp_sim = float(e.get("_similarity", 0.0))
                break
        top1 = results[0] if results else None
        top1_raw = raw[0] if raw else None
        top1_sim = float(top1_raw.get("_similarity", 0.0)) if top1_raw else 0.0
        top1_tr = top1.translation if top1 else ""

        expect_hit = bool(expected)
        decided_hit = len(results) > 0
        ok = (decided_hit == expect_hit)
        if ok:
            judge_ok += 1

        if expect_hit:
            hit_sims.append(exp_sim)
            # 路径分类:该查询若经别名层确定性命中(且命中的正是期望术语),记别名路径;
            # 否则语义检索路径(近似,依赖 RAG_MIN_SIMILARITY)。
            al = lookup_alias(query, "", "")
            is_alias = bool(al) and al[0] == expected
            if is_alias:
                alias_total += 1
                if rank == 1:
                    alias_top1 += 1
            else:
                sem_total += 1
                if rank == 1:
                    sem_top1 += 1
            if rank == 1:
                r1 += 1; r5 += 1; mrr_sum += 1.0
            elif rank > 0:
                r5 += 1; mrr_sum += 1.0 / rank
        else:
            miss_max_sims.append(top1_sim)
            if decided_hit:
                fp_rows.append((query, top1.term if top1 else "", top1_tr, top1_sim))

        mark = "PASS" if ok else "FAIL"
        print(f"[{mark}] {query!r:<46} rank={rank or '-':<2} "
              f"top1={top1.term if top1 else '∅'}({top1_sim:.2f}) "
              f"| {note} | {dt*1000:.0f}ms")
        if not ok and expect_hit:
            fail_rows.append((query, expected, top1.term if top1 else "", exp_sim))

    recall1 = r1 / n_hit * 100 if n_hit else 0.0
    recall5 = r5 / n_hit * 100 if n_hit else 0.0
    mrr = mrr_sum / n_hit if n_hit else 0.0
    judge_acc = judge_ok / n * 100 if n else 0.0

    print("=" * 100)
    print(f"Recall@1 = {r1}/{n_hit} = {recall1:.1f}%")
    print(f"Recall@5 = {r5}/{n_hit} = {recall5:.1f}%")
    print(f"MRR      = {mrr:.3f}")
    print(f"命中判定准确率 = {judge_ok}/{n} = {judge_acc:.1f}%")

    print("-" * 100)
    print("路径分类统计(按 search_rag 实际命中路径):")
    if alias_total:
        print(f"  别名路径(查字典·确定性): {alias_total}/{n_hit} 条查询 "
              f"→ top1 命中 {alias_top1}(召回率 {alias_top1/alias_total*100:.1f}%)")
    else:
        print("  别名路径: 0 条查询(测试集无别名可命中的表面形式)")
    if sem_total:
        print(f"  语义路径(近似·≥{config.RAG_MIN_SIMILARITY:.2f}): {sem_total}/{n_hit} 条查询 "
              f"→ top1 命中 {sem_top1}(召回率 {sem_top1/sem_total*100:.1f}%)")
    else:
        print(f"  语义路径: 0 条查询(所有查询均被别名层覆盖,语义路径未被本次测试集触及)")
    print(f"  合计: {n_hit} 条查询 → top1 命中 {r1}(Recall@1 = {recall1:.1f}%)")
    print("  ── 说明:别名路径是「查字典」,100% 准确但只能覆盖词库已有形式;"
          "语义路径是「近似」,处理词库未精确覆盖的查询(改写/库外),能力见下方阈值敏感性表")

    print("-" * 100)
    print("阈值敏感性(0.60~0.85,仅语义能力评估;别名命中的查询不受阈值影响):")
    for t in [0.60, 0.65, 0.70, 0.75, 0.80, 0.85]:
        keep = sum(1 for s in hit_sims if s >= t)
        fp = sum(1 for s in miss_max_sims if s >= t)
        print(f"  阈值 ≥ {t:.2f}: 命中保留 {keep}/{n_hit} "
              f"(召回率 {keep/n_hit*100 if n_hit else 0:.0f}%)  |  负例误判 {fp}/{n-n_hit}")
    if hit_sims:
        srt = sorted(hit_sims)
        print(f"  命中查询相似度: min={srt[0]:.2f} max={srt[-1]:.2f} median={srt[n_hit//2]:.2f}")

    if fail_rows:
        print("-" * 100)
        print("命中类未通过项:")
        for query, expected, top1, sim in fail_rows:
            print(f"  {query!r}: 期望 {expected!r} / 实际 top1 {top1!r} / 期望相似度 {sim:.2f}")
    if fp_rows:
        print("-" * 100)
        print("负例误报项(search_rag 返回了非空):")
        for query, top1, tr, sim in fp_rows:
            print(f"  {query!r}: → {top1!r} 译={tr!r} sim={sim:.2f}")

    # 文章覆盖度:200 条中在文章出现多少
    print("-" * 100)
    print(f"文章覆盖度:出现词库术语 {len(present)}/{len(glossary)} "
          f"({len(present)/len(glossary)*100:.1f}%)")
    print("出现术语:", " | ".join(t["term"] for t in present))


if __name__ == "__main__":
    main()

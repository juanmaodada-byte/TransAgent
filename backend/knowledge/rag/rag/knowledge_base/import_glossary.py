"""导入用户术语表到 RAG 术语库(内置库 / 用户上传两用)。

用法:
    python knowledge_base/import_glossary.py your_terms.csv [--user default] [--rebuild]

支持:
  - 格式:CSV(UTF-8/GBK 自动识别)、xlsx(需 openpyxl)、Markdown 表格
  - 列名自动识别(表头含以下任一即匹配):
      term      : term | english | en | 英文 | 原文 | 术语
      translation: translation | 译文 | 中文 | zh | 释义 | 译法
      domain    : domain | 领域 | 分类 | 子领域      (可选)
      action    : action | 译不译 | 是否翻译 | 不译  (可选)
      confidence: confidence | 置信度 | 确信度       (可选)
  - 无领域术语 → domain=""(全局可查,配合 search_rag 的 $or 兜底)
  - 同词同域冲突:报告冲突行,不静默覆盖
  - --rebuild:先清空 terms 集合再导入(用于把官方术语设为内置库);不加则追加
"""
import argparse
import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from contracts import TermEntry                      # noqa: E402
from knowledge_base import config, write_rag         # noqa: E402
from knowledge_base.glossary_split import split_combined_term  # noqa: E402
from knowledge_base.rag_terms import _client         # noqa: E402

# ── 列名别名(小写、去空格后匹配)────────────────────────────
_COL_ALIASES = {
    "term":        {"term", "english", "en", "英文", "原文", "术语", "terms"},
    "translation": {"translation", "译文", "中文", "zh", "释义", "译法", "translate"},
    "domain":      {"domain", "领域", "分类", "子领域", "category"},
    "action":      {"action", "译不译", "是否翻译", "不译", "translate?no"},
    "confidence":  {"confidence", "置信度", "确信度", "确认度"},
}

_MD_ALIASES = {"md", "markdown", "markdown.md"}
_ACTION_MAP = {"translate": "translate", "译": "translate", "翻译": "translate",
               "notranslate": "notranslate", "不译": "notranslate", "保留": "notranslate"}


def _norm(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def _detect_col(headers: list[str], kind: str) -> int | None:
    for i, h in enumerate(headers):
        if _norm(h) in _COL_ALIASES[kind]:
            return i
    return None


def _detect_encoding(path: Path) -> str:
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            path.read_text(encoding=enc, errors="strict")
            return enc
        except UnicodeDecodeError:
            continue
    return "utf-8"


def _read_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("xlsx 需要 openpyxl,请先 pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        return [[("" if c is None else str(c)).strip() for c in row]
                for row in ws.iter_rows(values_only=True) if any(row)]
    if suffix == ".csv" or suffix == ".txt":
        with open(path, encoding=_detect_encoding(path), newline="") as f:
            return [r for r in csv.reader(f) if any(c.strip() for c in r)]
    if suffix in {".md", ".markdown"}:
        out = []
        for line in path.read_text(encoding="utf-8-sig", errors="replace").splitlines():
            line = line.strip()
            if line.startswith("|") and line.endswith("|") and not set(line[1:-1]) <= {"-", " ", "|"}:
                out.append([c.strip() for c in line.strip("|").split("|")])
        return out
    raise RuntimeError(f"不支持的格式 {suffix}。请用 CSV / xlsx / Markdown。")


def ingest_glossary(path: str | Path, user_id: str = "", domain_default: str = "",
                    source: str = "用户术语表") -> tuple[list[TermEntry], list[str]]:
    """解析术语表文件 → (TermEntry 列表, 错误/冲突行说明)。不写库,可被 A 后端复用。"""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"文件不存在:{p}")
    rows = _read_rows(p)
    if not rows:
        return [], ["文件为空"]

    headers = [_norm(h) for h in rows[0]]
    col = {k: _detect_col(rows[0], k) for k in _COL_ALIASES}
    if col["term"] is None or col["translation"] is None:
        raise ValueError(
            f"未找到 term/translation 列。表头:{rows[0]}。"
            "可用的 term 列名:term/english/en/英文/原文/术语;translation 列名:translation/译文/中文/zh/释义")

    entries: list[TermEntry] = []
    errors: list[str] = []
    seen: dict[tuple[str, str], tuple[str, int]] = {}  # (domain, term.lower()) -> (译文, 行号)
    for ln, r in enumerate(rows[1:], start=2):
        term = r[col["term"]].strip() if col["term"] < len(r) else ""
        translation = r[col["translation"]].strip() if col["translation"] < len(r) else ""
        if not term or not translation:
            errors.append(f"第{ln}行缺 term 或 translation,已跳过:{r}")
            continue
        domain = (r[col["domain"]].strip() if col["domain"] is not None and col["domain"] < len(r)
                  else domain_default)
        action = (r[col["action"]].strip().lower() if col["action"] is not None and col["action"] < len(r)
                  else "")
        confidence = (r[col["confidence"]].strip().lower() if col["confidence"] is not None and col["confidence"] < len(r)
                      else "high")
        key = (domain, term.lower())
        if key in seen:                       # ── 原行级重复检测(对原词条名,含缝合怪) ──
            errors.append(f"第{ln}行与第{seen[key][1]}行冲突:同词同域({domain!r}|{term!r})译文不同"
                          f"({seen[key][0]!r} vs {translation!r}),后者已跳过")
            continue
        seen[key] = (translation, ln)         # 登记原词条键:后续重复的缝合怪原词仍能检出冲突

        # ── 拆分缝合怪词条:每个拆分产物独立参与 (domain, term.lower()) 去重 ──
        # 未拆分的普通词条 skey==key 已登记(自己撞自己),直接通过;
        # 拆分产物 skey!=key,与库内已有词条撞名时报冲突(先到先得)。
        for sterm, strans in split_combined_term(term, translation):
            skey = (domain, sterm.lower())
            if skey != key and skey in seen:
                errors.append(
                    f"第{ln}行拆分产物与第{seen[skey][1]}行冲突:同词同域"
                    f"({domain!r}|{sterm!r})译文不同({seen[skey][0]!r} vs {strans!r}),"
                    f"该拆分产物已跳过")
                continue
            seen[skey] = (strans, ln)
            entries.append(TermEntry(
                term=sterm,
                translation=strans,
                domain=domain,
                confidence=confidence if confidence in {"high", "medium", "low"} else "high",
                action=_ACTION_MAP.get(action, "translate"),
                source=source,
            ))
    return entries, errors


def main() -> None:
    ap = argparse.ArgumentParser(description="导入术语表到 RAG")
    ap.add_argument("file", help="CSV/xlsx/Markdown 路径")
    ap.add_argument("--user", default=config.DEFAULT_USER, help="user_id(默认 default)")
    ap.add_argument("--source", default="用户术语表", help="来源标注(内置库建议:官方术语库)")
    ap.add_argument("--rebuild", action="store_true", help="先清空 terms 集合再导入(设为内置库)")
    args = ap.parse_args()

    entries, errors = ingest_glossary(args.file, user_id=args.user, source=args.source)
    if errors:
        print(f"⚠ 解析提示 {len(errors)} 条:")
        for e in errors[:10]:
            print("   ", e)
        if len(errors) > 10:
            print(f"    … 共 {len(errors)} 条")
    if not entries:
        print("没有可导入的有效术语,退出")
        return

    if args.rebuild:
        from knowledge_base.rag_terms import reset_collection
        reset_collection(args.user)
        print(f"已清空 {config.TERMS_COLLECTION} 集合 + 别名表(rebuild)")

    ids = write_rag(entries, user_id=args.user)
    print(f"✅ 导入 {len(entries)} 条术语({Path(args.file).name} → user_id={args.user!r})")
    print(f"   带领域 {sum(1 for e in entries if e.domain)} 条 / 无领域(全局) {sum(1 for e in entries if not e.domain)} 条")
    print(f"   同词同域冲突 {len(errors)} 条(见上方提示)")


if __name__ == "__main__":
    main()
